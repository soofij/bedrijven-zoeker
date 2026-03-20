import streamlit as st
import requests
import pandas as pd
import time
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

st.set_page_config(page_title="Bedrijven Zoeker", page_icon="🔍")
st.title("Bedrijven Zoeker")
st.write("Zoek bedrijven op thema en regio en download de resultaten als Excel.")

api_key = st.text_input("Serper API key", type="password")
steden_input = st.text_input("Steden (kommagescheiden)", "Arnhem, Apeldoorn, Zutphen, Deventer, Doetinchem")
zoektermen_input = st.text_input("Zoektermen (kommagescheiden)", "duurzaamheid, circulariteit, warmtetransitie, klimaatadvies")
max_resultaten = st.slider("Max resultaten per zoekopdracht", 5, 50, 10)

filter_urls = [
    ".nl/gemeente", "gemeente.nl", "provincie.nl",
    "klimaatbeheersing", "klimaattechniek", "installatie",
    "bouwbedrijf", "acel.nl", "despaan.nl", "ktd-", "kso-",
    "bergevoet.nl", "oude-ijsselstreek.nl", "/gebruikte-bouwmaterialen",
    "/voorbeelden-van-", "/duurzame-ondernemers",
    "arnhem.nl/alle-onderwerpen", "deventer.nl/", "apeldoorn.nl/",
    "zutphen.nl/", "doetinchem.nl/", "gelderland.nl/"
]

filter_woorden = [
    "wat wij doen", "home", "nieuws", "over ons", "contact",
    "blog", "vacatures", "producten", "diensten", "welkom",
    "bedrijven", "zoeken", "resultaten", "pagina", "artikel",
    "wikipedia", "linkedin.com/posts", "youtube", "facebook"
]

filter_titels = [
    "installatie", "klimaattechniek", "klimaatbeheersing",
    "bouwbedrijf", "gemeente", "provincie", "techniek b.v",
    "loodgieter", "elektricien", "aannemer",
    "gebruikte bouwmaterialen", "voorbeelden van"
]

headers = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}

def is_geen_bedrijf(titel, link):
    titel_lower = titel.lower()
    link_lower = link.lower()
    for woord in filter_woorden:
        if woord in titel_lower or woord in link_lower:
            return True
    for woord in filter_urls:
        if woord in link_lower:
            return True
    for woord in filter_titels:
        if woord in titel_lower:
            return True
    return False

def schoon_naam_op(titel):
    for scheidingsteken in [" | ", " - ", " :: ", " — "]:
        if scheidingsteken in titel:
            delen = titel.split(scheidingsteken)
            delen = [d.strip() for d in delen if len(d.strip()) > 3]
            if delen:
                titel = min(delen, key=len)
    return titel.strip()

def haal_omschrijving_op(url):
    try:
        response = requests.get(url, headers=headers, timeout=8)
        content_type = response.headers.get("Content-Type", "")
        if "text/html" not in content_type:
            return ""
        soup = BeautifulSoup(response.text, "html.parser")
        for tag in soup(["nav", "footer", "script", "style", "header"]):
            tag.decompose()
        teksten = []
        for p in soup.find_all("p"):
            tekst = p.get_text(strip=True)
            if len(tekst) > 60:
                teksten.append(tekst)
            if len(teksten) >= 2:
                break
        if teksten:
            omschrijving = " ".join(teksten)
            if len(omschrijving) > 300:
                afgekapt = omschrijving[:300]
                laatste_punt = max(afgekapt.rfind("."), afgekapt.rfind("!"), afgekapt.rfind("?"))
                if laatste_punt > 100:
                    omschrijving = afgekapt[:laatste_punt + 1]
                else:
                    omschrijving = afgekapt + "..."
            return omschrijving
        return ""
    except:
        return ""

def zoek_bedrijven(zoekterm, stad, api_key, max_resultaten):
    query = f"{zoekterm} bedrijf {stad}"
    url = "https://google.serper.dev/search"
    headers_api = {
        "X-API-KEY": api_key,
        "Content-Type": "application/json"
    }
    payload = {"q": query, "hl": "nl", "gl": "nl", "num": max_resultaten}
    try:
        response = requests.post(url, headers=headers_api, json=payload, timeout=10)
        data = response.json()
        resultaten = []
        for item in data.get("organic", []):
            titel = schoon_naam_op(item.get("title", ""))
            link = item.get("link", "")
            snippet = item.get("snippet", "")
            if titel and link and not is_geen_bedrijf(titel, link):
                omschrijving = haal_omschrijving_op(link)
                if not omschrijving:
                    omschrijving = snippet
                resultaten.append({
                    "Bedrijfsnaam": titel,
                    "Website": link,
                    "Omschrijving": omschrijving,
                    "Zoekterm": zoekterm,
                    "Stad": stad
                })
        return resultaten
    except Exception as e:
        return []

if st.button("Zoeken en Excel downloaden"):
    if not api_key:
        st.error("Vul je Serper API key in!")
    else:
        steden = [s.strip() for s in steden_input.split(",")]
        zoektermen = [z.strip() for z in zoektermen_input.split(",")]
        alle_resultaten = []
        totaal = len(steden) * len(zoektermen)
        voortgang = st.progress(0)
        status = st.empty()
        stap = 0

        for stad in steden:
            for zoekterm in zoektermen:
                status.text(f"Zoeken: {zoekterm} in {stad}...")
                resultaten = zoek_bedrijven(zoekterm, stad, api_key, max_resultaten)
                alle_resultaten.extend(resultaten)
                stap += 1
                voortgang.progress(stap / totaal)
                time.sleep(1)

        df = pd.DataFrame(alle_resultaten)
        df = df.drop_duplicates(subset=["Website"])
        df = df.reset_index(drop=True)

        output = io.BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)

        wb = load_workbook(output)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            cel = row[1]
            if cel.value and str(cel.value).startswith("http"):
                cel.hyperlink = cel.value
                cel.font = Font(color="0000FF", underline="single")

        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for cel in ws[1]:
            cel.fill = header_fill
            cel.font = header_font
            cel.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        lichtgrijs = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        wit = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            for cel in row:
                cel.fill = lichtgrijs if i % 2 == 0 else wit
                cel.alignment = Alignment(vertical="center", wrap_text=False)
            ws.row_dimensions[i].height = 18

        for column in ws.columns:
            max_breedte = 0
            for cel in column:
                if cel.value:
                    max_breedte = max(max_breedte, len(str(cel.value)))
            ws.column_dimensions[column[0].column_letter].width = min(max_breedte + 2, 50)

        final_output = io.BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        status.text(f"Klaar! {len(df)} bedrijven gevonden.")
        voortgang.progress(1.0)

        st.success(f"{len(df)} bedrijven gevonden!")
        st.download_button(
            label="Download Excel",
            data=final_output,
            file_name="bedrijven_resultaten.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
